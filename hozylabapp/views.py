import io
import os
from datetime import datetime

import pandas as pd
import pytz
from django.http import HttpResponse
from openpyxl import Workbook
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render

# Create your views here.
from django.views.generic import DetailView, ListView
from openpyxl.writer.excel import save_virtual_workbook

from assetmasterapp.models import Asset
from cryptoapp.models import Crypto, CryptoTransaction
from diamond_goose_mk6 import settings
from diamond_goose_mk6.settings import base
from equityapp.models import Equity, EquityTransaction
from exchangeapp.models import ForeignCurrency, ForeignCurrencyTransaction
from guardianapp.models import Guardian, GuardianTransaction
from hozylabapp.models import TempTransaction
from pensionapp.models import PensionAsset, PensionAssetTransaction
from reitsapp.models import Reits, ReitsTransaction


def lab_home_view(request):
    return render(request, 'hozylabapp/lab_home.html')


def temptransaction_home_view(request):
    return render(request, 'hozylabapp/temptransaction_home.html')


class TempTransactionListView(ListView):
    model = TempTransaction
    context_object_name = 'temp_transaction_list'
    template_name = 'hozylabapp/temptransaction_list.html'

    def get_context_data(self, **kwargs):
        context = super(TempTransactionListView, self).get_context_data(**kwargs)

        return context


# class TempTransactionDetailView(DetailView):
#     model = TempTransaction
#     # context_object_name = 'target_foreign_currency'
#     # form_class = ForeignCurrencyTransactionCreationForm
#     template_name = 'hozylabapp/temptransaction_detail.html'
#
#     def get_context_data(self, **kwargs):
#         context = super(TempTransactionDetailView, self).get_context_data(**kwargs)
#
#         return context


def read_excel(request):
    # transaction_file = request.FILES['transaction_file']
    # in_media_directory = 'convert_transaction_excel'
    #
    # company = request.POST['company']
    # in_media_directory += '/'
    # in_media_directory += company
    # in_media_directory += '/'
    #
    # save_dir = os.path.join(base.MEDIA_ROOT, in_media_directory)
    # fs = FileSystemStorage(location=save_dir)
    # file_name = fs.save(transaction_file.name, transaction_file)
    # new_dir = in_media_directory + file_name
    #
    # uploaded_file_url = fs.url(new_dir)
    # excel_file = uploaded_file_url

    excel_transaction_data = pd.read_excel(request.FILES['transaction_file'], sheet_name=0, header=[0, 1])
    return excel_transaction_data.fillna('')


def asset_type_determinate(ticker):
    try:
        queryset_registerd_asset = Asset.objects.get(ticker=ticker)
        if queryset_registerd_asset:
            asset_type = queryset_registerd_asset.asset_type
    except:
        import yfinance as yf
        new_asset = yf.Ticker(ticker).get_info()
        if new_asset:
            print(new_asset)
            print('New Asset Creation : ', ticker)
            if new_asset['longName']:
                asset_name = new_asset['longName']
            else:
                asset_name = new_asset['shortName']
            print(asset_name)
            print(new_asset['currency'])
            print(new_asset['previousClose'])
            obj = Asset.objects.create(
                    ticker=ticker,
                    name=asset_name,
                    currency=new_asset['currency'],
                    current_price=new_asset['previousClose'],
                    pension_non_risk_asset_flag=False,
            )
            obj.save()
            print('-> Creation Completed.')

        asset_type = 'UNDEFINED'

    return asset_type


def upload_mass_transaction(request):
    try:
        if request.method == 'POST' and request.FILES['transaction_file']:
            db_frame = read_excel(request)
    except Exception as identifier:
        print('upload_mass_transaction: excel_import', identifier)

    try:
        transaction_object_list = []

        for row in db_frame.itertuples():
            upload_format = {}
            upload_format['data_source'] = row[1]
            upload_format['asset_type'] = row[2]
            upload_format['transaction_type'] = row[3]
            upload_format['ticker'] = str(row[4]).split("'")[-1] if "'" in row[4] else row[4]
            upload_format['pension_type'] = row[5]
            upload_format['currency'] = row[6]
            upload_format['quantity'] = 0 if row[7] == '' else row[7]
            upload_format['price'] = 0 if row[8] == '' else row[8]
            upload_format['exchange_rate'] = 0 if row[9] == '' else row[9]
            upload_format['transaction_fee'] = 0 if row[10] == '' else row[10]
            upload_format['transaction_tax'] = 0 if row[11] == '' else row[11]
            upload_format['split_ratio_one_to_N'] = 0 if row[12] == '' else row[12]
            upload_format['transaction_date'] = row[13]

            if upload_format['data_source'] != 'data_source':
                tz = pytz.timezone('Asia/Seoul')
                transaction_date_raw = upload_format['transaction_date'].to_pydatetime()
                transaction_date_timezone = tz.localize(transaction_date_raw, is_dst=None).astimezone(pytz.utc)

                if upload_format['asset_type'] in ['EQUITY', 'CRYPTO', 'GUARDIAN', 'REITS', 'PENSION']:
                    try:
                        target_asset = Asset.objects.get(ticker=upload_format['ticker'])
                    except Exception as identifier:
                        print('Exception in calling target Asset model by ticker.', identifier, upload_format)
                        break
                    try:
                        if upload_format['asset_type'] == 'EQUITY':
                            target_my_asset = Equity.objects.get(owner=request.user,
                                                                 asset=target_asset.pk)
                            obj = EquityTransaction.objects.create(
                                equity=target_my_asset,
                                transaction_type=upload_format['transaction_type'],
                                quantity=upload_format['quantity'],
                                price=upload_format['price'],
                                transaction_fee=upload_format['transaction_fee'],
                                transaction_tax=upload_format['transaction_tax'],
                                transaction_date=transaction_date_timezone,
                                note='Excel Mass Upload '+upload_format['data_source'],
                            )

                        elif upload_format['asset_type'] == 'CRYPTO':
                            target_my_asset = Crypto.objects.get(owner=request.user,
                                                                 asset=target_asset.pk)
                            obj = CryptoTransaction.objects.create(
                                crypto=target_my_asset,
                                transaction_type=upload_format['transaction_type'],
                                quantity=upload_format['quantity'],
                                price=upload_format['price'],
                                transaction_fee=upload_format['transaction_fee'],
                                transaction_tax=upload_format['transaction_tax'],
                                transaction_date=transaction_date_timezone,
                                note='Excel Mass Upload '+upload_format['data_source'],
                            )

                        elif upload_format['asset_type'] == 'GUARDIAN':
                            target_my_asset = Guardian.objects.get(owner=request.user,
                                                                   asset=target_asset.pk)
                            obj = GuardianTransaction.objects.create(
                                guardian=target_my_asset,
                                transaction_type=upload_format['transaction_type'],
                                quantity=upload_format['quantity'],
                                price=upload_format['price'],
                                transaction_fee=upload_format['transaction_fee'],
                                transaction_tax=upload_format['transaction_tax'],
                                transaction_date=transaction_date_timezone,
                                note='Excel Mass Upload '+upload_format['data_source'],
                            )

                        elif upload_format['asset_type'] == 'REITS':
                            target_my_asset = Reits.objects.get(owner=request.user,
                                                                asset=target_asset.pk)
                            obj = ReitsTransaction.objects.create(
                                reits=target_my_asset,
                                transaction_type=upload_format['transaction_type'],
                                quantity=upload_format['quantity'],
                                price=upload_format['price'],
                                transaction_fee=upload_format['transaction_fee'],
                                transaction_tax=upload_format['transaction_tax'],
                                transaction_date=transaction_date_timezone,
                                note='Excel Mass Upload '+upload_format['data_source'],
                            )

                        elif upload_format['asset_type'] == 'PENSION':
                            target_my_asset = PensionAsset.objects.get(owner=request.user,
                                                                       asset=target_asset.pk)
                            obj = PensionAssetTransaction.objects.create(
                                pension_asset=target_my_asset,
                                transaction_type=upload_format['transaction_type'],
                                quantity=upload_format['quantity'],
                                price=upload_format['price'],
                                transaction_fee=upload_format['transaction_fee'],
                                transaction_tax=upload_format['transaction_tax'],
                                transaction_date=transaction_date_timezone,
                                note='Excel Mass Upload '+upload_format['data_source'],
                            )

                    except Exception as identifier:
                        print('Exception in calling target {} model by ticker.'.format(upload_format['asset_type']), identifier, upload_format)

                elif upload_format['asset_type'] == 'EXCHANGE':
                    try:
                        target_foreign_currency = ForeignCurrency.objects.get(currency=upload_format['currency'],
                                                                              owner=request.user)
                        obj = ForeignCurrencyTransaction.objects.create(
                            foreign_currency=target_foreign_currency,
                            transaction_type=upload_format['transaction_type'],
                            quantity=upload_format['quantity'],
                            exchange_rate=upload_format['exchange_rate'],
                            transaction_date=transaction_date_timezone,
                            note='Excel Mass Upload '+upload_format['data_source'],
                        )
                    except Exception as identifier:
                        print('Exception in calling target ForeignCurrency model by currency.', identifier, upload_format)

                else:
                    print('Unexpected Transaction exists : ', upload_format)

            transaction_object_list.append(obj)

        # Transaction COMMIT
        for transaction in transaction_object_list:
            try:
                transaction.save()
                print('Transaction Added :', transaction)
            except Exception as identifier:
                print('Exception in saving transaction.', identifier, transaction)
                continue

    except Exception as identifier:
        print('upload_mass_transaction: Transaction INSERT :', identifier)

    return render(request, 'hozylabapp/temptransaction_list.html')



def upload_excel_daeshin(request):
    try:
        if request.method == 'POST' and request.FILES['transaction_file']:
            db_frame = read_excel(request)
    except Exception as identifier:
        print('upload_dashin: excel_import', identifier)

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "transactions"

        output_row_count = 1

        # Header Insert
        worksheet.cell(row=output_row_count, column=1, value="data_source")
        worksheet.cell(row=output_row_count, column=2, value="asset_type")
        worksheet.cell(row=output_row_count, column=3, value="transaction_type")
        worksheet.cell(row=output_row_count, column=4, value="ticker")
        worksheet.cell(row=output_row_count, column=5, value="pension_type")
        worksheet.cell(row=output_row_count, column=6, value="currency")
        worksheet.cell(row=output_row_count, column=7, value="quantity")
        worksheet.cell(row=output_row_count, column=8, value="price")
        worksheet.cell(row=output_row_count, column=9, value="exchange_rate")
        worksheet.cell(row=output_row_count, column=10, value="transaction_fee")
        worksheet.cell(row=output_row_count, column=11, value="transaction_tax")
        worksheet.cell(row=output_row_count, column=12, value="split_ratio_one_to_N")
        worksheet.cell(row=output_row_count, column=13, value="transaction_date")
        worksheet.cell(row=output_row_count, column=14, value="applied_flag")
        worksheet.cell(row=output_row_count, column=15, value="applied_date")

        # Line Insert
        for row in db_frame.itertuples():
            if int(row[0]) % 2 == 0:
                daeshin_format = {}
                daeshin_format['거래일'] = row[1]
                daeshin_format['거래구분'] = row[2]
                daeshin_format['통화'] = row[3]
                daeshin_format['거래금액'] = row[4]
                daeshin_format['질권일'] = row[5]
                daeshin_format['입금환율'] = row[6]
                daeshin_format['종목코드'] = row[7]
                daeshin_format['수량'] = row[8]
                daeshin_format['유가잔고'] = row[9]
                daeshin_format['국내세'] = row[10]
                daeshin_format['제미납금'] = row[11]
                daeshin_format['외화결제금액'] = row[12]
                daeshin_format['거래상대명'] = row[13]

            else:
                daeshin_format['순번'] = row[1]
                daeshin_format['적요명'] = row[2]
                daeshin_format['환전'] = row[3]
                daeshin_format['환전금액'] = row[4]
                daeshin_format['상환금액'] = row[5]
                daeshin_format['출금환율'] = row[6]
                daeshin_format['종목명'] = row[7]
                daeshin_format['단가'] = row[8]
                daeshin_format['수수료'] = row[9]
                daeshin_format['현지세'] = row[10]
                daeshin_format['연체신용이자'] = row[11]
                daeshin_format['외화예수금'] = row[12]
                daeshin_format['원화예수금'] = row[13]

                # Variable Initialization
                asset_type = None
                transaction_type = None
                ticker = None
                pension_type = None
                currency = None
                quantity = 0
                price = 0
                exchange_rate = 0
                transaction_fee = 0
                transaction_tax = 0
                split_ratio_one_to_N = 0
                transaction_date = datetime.strptime(daeshin_format['거래일'], "%Y/%m/%d")
                valid_transaction_flag = False


                # Equity Transaction
                if daeshin_format['거래구분'] == '해외증권장내매매':
                    ticker = daeshin_format['종목코드']
                    if ticker == 'BRK.B':
                        ticker = 'BRK-B'
                    asset_type = asset_type_determinate(ticker)
                    currency = daeshin_format['통화']
                    quantity = daeshin_format['수량']
                    price = daeshin_format['단가']
                    transaction_fee = daeshin_format['수수료']
                    transaction_tax = daeshin_format['현지세']
                    if daeshin_format['적요명'] == '현금매수':
                        transaction_type = 'BUY'
                        valid_transaction_flag = True
                    elif daeshin_format['적요명'] == '현금매도':
                        transaction_type = 'SELL'
                        valid_transaction_flag = True

                # Dividend
                if daeshin_format['적요명'] == '배당금' and daeshin_format['거래구분'] == '입금':
                    transaction_type = 'DIVIDEND'
                    ticker = daeshin_format['종목코드']
                    asset_type = asset_type_determinate(ticker)
                    currency = daeshin_format['통화']
                    quantity = 1
                    price = daeshin_format['거래금액']
                    transaction_tax = daeshin_format['현지세']
                    valid_transaction_flag = True

                # Split
                if daeshin_format['적요명'] == '액면교체':
                    transaction_type = 'SPLIT'
                    ticker = daeshin_format['종목코드']
                    asset_type = asset_type_determinate(ticker)
                    currency = daeshin_format['통화']
                    split_ratio_one_to_N = 1
                    valid_transaction_flag = True

                # Foreign Exchange Sell
                if daeshin_format['적요명'] == '외화매도환전' and daeshin_format['거래구분'] == '출금':
                    # print('외화매도환전', daeshin_format)
                    asset_type = 'EXCHANGE'
                    transaction_type = 'SELL'
                    currency = daeshin_format['통화']
                    quantity = daeshin_format['거래금액']
                    exchange_rate = daeshin_format['출금환율']
                    valid_transaction_flag = True

                # Foreign Exchange Buy
                if daeshin_format['적요명'] == '외화매수환전' and daeshin_format['거래구분'] == '입금':
                    # print('외화매수환전', daeshin_format)
                    asset_type = 'EXCHANGE'
                    transaction_type = 'BUY'
                    currency = daeshin_format['환전']
                    quantity = daeshin_format['환전금액']
                    exchange_rate = daeshin_format['입금환율']
                    valid_transaction_flag = True

                # Excel Export Target Transactions
                if valid_transaction_flag:
                    output_row_count += 1
                    worksheet.cell(row=output_row_count, column=1, value="DAESHIN")     # data_source
                    worksheet.cell(row=output_row_count, column=2, value=asset_type)
                    worksheet.cell(row=output_row_count, column=3, value=transaction_type)
                    worksheet.cell(row=output_row_count, column=4, value=ticker)
                    worksheet.cell(row=output_row_count, column=5, value=pension_type)
                    worksheet.cell(row=output_row_count, column=6, value=currency)
                    worksheet.cell(row=output_row_count, column=7, value=quantity)
                    worksheet.cell(row=output_row_count, column=8, value=price)
                    worksheet.cell(row=output_row_count, column=9, value=exchange_rate)
                    worksheet.cell(row=output_row_count, column=10, value=transaction_fee)
                    worksheet.cell(row=output_row_count, column=11, value=transaction_tax)
                    worksheet.cell(row=output_row_count, column=12, value=split_ratio_one_to_N)
                    worksheet.cell(row=output_row_count, column=13, value=transaction_date)

        response = HttpResponse(content=save_virtual_workbook(workbook),
                                content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=transaction_daeshin.xlsx'
        return response

    except Exception as identifier:
        print('upload_dashin: excel_export', identifier)


def upload_excel_hankook(request):

    def comma_remover(string_number):
        result_list = []
        for i in string_number:
            if i != ',':
                result_list.append(i)
        if len(result_list) == 0:
            result_list.append('0')
        return float(''.join(result_list))

    try:
        if request.method == 'POST' and request.FILES['transaction_file']:
            db_frame = read_excel(request)
    except Exception as identifier:
        print('upload_hankook: excel_import', identifier)

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "transactions"

        output_row_count = 1

        # Header Insert
        worksheet.cell(row=output_row_count, column=1, value="data_source")
        worksheet.cell(row=output_row_count, column=2, value="asset_type")
        worksheet.cell(row=output_row_count, column=3, value="transaction_type")
        worksheet.cell(row=output_row_count, column=4, value="ticker")
        worksheet.cell(row=output_row_count, column=5, value="pension_type")
        worksheet.cell(row=output_row_count, column=6, value="currency")
        worksheet.cell(row=output_row_count, column=7, value="quantity")
        worksheet.cell(row=output_row_count, column=8, value="price")
        worksheet.cell(row=output_row_count, column=9, value="exchange_rate")
        worksheet.cell(row=output_row_count, column=10, value="transaction_fee")
        worksheet.cell(row=output_row_count, column=11, value="transaction_tax")
        worksheet.cell(row=output_row_count, column=12, value="split_ratio_one_to_N")
        worksheet.cell(row=output_row_count, column=13, value="transaction_date")
        worksheet.cell(row=output_row_count, column=14, value="applied_flag")
        worksheet.cell(row=output_row_count, column=15, value="applied_date")

        # Line Insert
        for row in db_frame.itertuples():
            if int(row[0]) % 2 == 0:
                hankook_format = {}
                hankook_format['거래일'] = row[1]
                hankook_format['종목명'] = row[2]
                hankook_format['거래수량'] = comma_remover(row[3])
                hankook_format['환율'] = comma_remover(row[4])
                hankook_format['거래금액'] = comma_remover(row[5])
                hankook_format['수수료'] = comma_remover(row[6])
                hankook_format['유가잔고'] = comma_remover(row[7])
                hankook_format['잔액'] = comma_remover(row[8])
                hankook_format['상대계좌'] = row[9]

            else:
                hankook_format['거래종류'] = row[1]
                hankook_format['잔고번호'] = row[2]
                hankook_format['거래단가'] = comma_remover(row[3])
                hankook_format['외화잔액'] = comma_remover(row[4])
                hankook_format['정산금액'] = comma_remover(row[5])
                hankook_format['거래세'] = comma_remover(row[6])
                hankook_format['세금'] = comma_remover(row[7])
                hankook_format['부가세'] = comma_remover(row[8])
                hankook_format['접속매체'] = row[9]

                # Variable Initialization
                asset_type = None
                transaction_type = None
                ticker = None
                pension_type = None
                currency = None
                quantity = 0
                price = 0
                exchange_rate = 0
                transaction_fee = 0
                transaction_tax = 0
                split_ratio_one_to_N = 0
                transaction_date = datetime.strptime(hankook_format['거래일'], "%Y.%m.%d")
                valid_transaction_flag = False


                # Equity Transaction
                if '삼성전자' in hankook_format['종목명']:
                    if hankook_format['종목명'] == '삼성전자' or '보통주' in hankook_format['종목명']:
                        ticker = "'005930"
                    else:
                        ticker = "'005935"

                    asset_type = asset_type_determinate(ticker.split("'")[-1])
                    currency = 'KRW'

                    if hankook_format['거래종류'] == 'Smart+거래소주식매수':
                        transaction_type = 'BUY'
                        quantity = hankook_format['거래수량']
                        price = hankook_format['거래단가']
                        transaction_fee = hankook_format['수수료']
                        transaction_tax = hankook_format['거래세']
                        valid_transaction_flag = True
                    elif hankook_format['거래종류'] == 'Smart+거래소주식매도':
                        transaction_type = 'SELL'
                        quantity = hankook_format['거래수량']
                        price = hankook_format['거래단가']
                        transaction_fee = hankook_format['수수료']
                        transaction_tax = hankook_format['거래세']
                        valid_transaction_flag = True
                    elif hankook_format['거래종류'] == '배당금입금':
                        transaction_type = 'DIVIDEND'
                        quantity = 1
                        price = hankook_format['거래금액']
                        transaction_fee = hankook_format['수수료']
                        transaction_tax = hankook_format['거래세']
                        valid_transaction_flag = True

                # Excel Export Target Transactions
                if valid_transaction_flag:
                    output_row_count += 1
                    worksheet.cell(row=output_row_count, column=1, value="HANKOOK")     # data_source
                    worksheet.cell(row=output_row_count, column=2, value=asset_type)
                    worksheet.cell(row=output_row_count, column=3, value=transaction_type)
                    worksheet.cell(row=output_row_count, column=4, value=ticker)
                    worksheet.cell(row=output_row_count, column=5, value=pension_type)
                    worksheet.cell(row=output_row_count, column=6, value=currency)
                    worksheet.cell(row=output_row_count, column=7, value=quantity)
                    worksheet.cell(row=output_row_count, column=8, value=price)
                    worksheet.cell(row=output_row_count, column=9, value=exchange_rate)
                    worksheet.cell(row=output_row_count, column=10, value=transaction_fee)
                    worksheet.cell(row=output_row_count, column=11, value=transaction_tax)
                    worksheet.cell(row=output_row_count, column=12, value=split_ratio_one_to_N)
                    worksheet.cell(row=output_row_count, column=13, value=transaction_date)

        response = HttpResponse(content=save_virtual_workbook(workbook),
                                content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=transaction_hankook.xlsx'
        return response

    except Exception as identifier:
        print('upload_dashin: excel_export', identifier)